from adze_modeler.modelpaths import ModelDir
import requests
import openpyxl as oxl
import json

ModelDir.set_base(__file__)
file_ = str(ModelDir.BASE / "transformer.xlsx")

wb = oxl.load_workbook(filename=file_)
ws = wb.worksheets[0]

data = {
    "simulation": {"type": "short_circuit_impedance",
                   "js": ws['B17'].value,
                   "jp": ws['B18'].value},
    "model": {
        "w1": ws['B2'].value, 
        "h1": ws['B3'].value, 
        "r1": ws['B4'].value, 
        "z1": ws['B5'].value, 
        "w2": ws['B7'].value, 
        "h2": ws['B8'].value, 
        "r2": ws['B9'].value, 
        "z2": ws['B10'].value, 
        "w3": ws['B12'].value, 
        "h3": ws['B13'].value, 
        "r3": ws['B14'].value,
        "z3": ws['B15'].value
    },
}
url = 'http://127.0.0.1:5000/process'
res = requests.post(url, data=json.dumps(data)).json().get('res')
print(res)


ws["B20"] = round(res['Xlv']*100, 3)
ws["B21"] = round(res['Xhv']*100, 3)
wb.save(filename=file_)
